import openpyxl
import requests


def criar_planilha(file_name: str) -> object:
    url_filename = f'C:/Users/julia/PycharmProjects/botcep/botdesafio/{file_name}.xlsx'
    book = openpyxl.Workbook()
    book.active.title = 'Cep'
    book.create_sheet('Modelo')
    cep_page = book['Cep']
    modelo_page = book['Modelo']
    cep_page.append(['CEP'])
    modelo_page.append(['CEP', 'Logradouro', 'Bairro', 'Localidade', 'UF', 'DDD'])

    book.save(url_filename)


def atualizar_tabela1(filename, sheet_name, column_name, content_list):
    list_items = content_list
    file_name = filename
    url_filename = f'C:/Users/julia/PycharmProjects/botcep/botdesafio/{file_name}.xlsx'
    book = openpyxl.load_workbook(url_filename)
    pagina_para_atualizar = book[sheet_name]
    titulos_colunas = {'CEP': 'A'}
    for i in range(0, len(list_items)):
        index = f'{titulos_colunas[column_name]}{i + 2}'
        pagina_para_atualizar[index] = list_items[i]
    book.save(url_filename)


def consumir_api(file_name_test):
    url_filename = f'C:/Users/julia/PycharmProjects/botcep/botdesafio/{file_name_test}.xlsx'
    url = 'https://viacep.com.br/ws/{}/json/'
    workbook = openpyxl.load_workbook(url_filename)

    worksheet_ceps = workbook['Cep']
    worksheet_resultados = workbook['Modelo']

    # Definir linha inicial para inserir os dados na planilha de resultados
    linha_resultados = 2

    for linha in worksheet_ceps.iter_rows(min_row=2, values_only=True):
        # Obter o CEP da linha atual
        cep1 = linha[0]

            # Fazer a requisição GET para a API do ViaCEP com o CEP atual
        response = requests.get(url.format(cep1))
        response.raise_for_status()
        data = response.json()

        assert isinstance(worksheet_resultados, object)
        worksheet_resultados.cell(row=linha_resultados, column=1).value = data['cep']
        worksheet_resultados.cell(row=linha_resultados, column=2).value = data['logradouro']
        worksheet_resultados.cell(row=linha_resultados, column=3).value = data['bairro']
        worksheet_resultados.cell(row=linha_resultados, column=4).value = data['localidade']
        worksheet_resultados.cell(row=linha_resultados, column=5).value = data['uf']
        worksheet_resultados.cell(row=linha_resultados, column=6).value = data['ddd']

        # Incrementar a linha para inserir os dados na próxima linha da planilha de resultados
        linha_resultados += 1

        # Salvar arquivo Excel com os resultados
        workbook.save(url_filename)






